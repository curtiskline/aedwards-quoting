#!/usr/bin/env python3
"""Seed product_catalog from Chip's provided spreadsheets."""

from __future__ import annotations

import argparse
from pathlib import Path

from app import create_app
from app.extensions import db
from app.models import ProductCatalog

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required. Install dependencies before running this script.") from exc


DEFAULT_PART_DESCRIPTIONS = Path("/tmp/chip-docs/Part Descriptions.xlsx")
DEFAULT_BOOK1 = Path("/tmp/chip-docs/Book1.xlsx")

CATEGORY_HEADERS = {
    "compression sleeves",
    "girth weld",
    "bags",
    "sleeves",
    "pipe jacks",
    "backing strips",
    "omegawrap",
    "accessories",
}

OMEGAWRAP_SKUS = {
    "Carbon",
    "E-Glass",
    "Isolation Wrap",
    "Magnum",
    "Resin",
    "Putty",
    "Magnet Set",
    "Porcupine Roller",
    "Accessory Kit",
    "Plastic Wrap",
}


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def classify_type(part_number: str) -> str | None:
    """Map a part number to a product Type slug (ProductType.name).

    Returns None for identifiers with no agreed type (the old PIPE_JACK and
    OTHER buckets), matching the family->type migration: those rows are left
    type-unset for a human to triage rather than guessed.
    """
    normalized = part_number.strip()
    upper = normalized.upper()
    lower = normalized.lower()

    if upper.startswith("S-"):
        return "sleeve"
    if upper.startswith("G-"):
        return "girth_weld"
    if upper.startswith("GTW ") or lower == "soft set":
        return "bag"
    if upper.startswith("PJ-"):
        return None  # was PIPE_JACK — no agreed type yet
    if lower == "backing strip":
        return "accessory"
    if normalized in OMEGAWRAP_SKUS:
        return "composite"
    if upper.startswith("CS-") or lower.startswith("compression sleeve"):
        return "compression"
    return None  # was OTHER — catch-all, leave untyped for triage


def looks_like_book1_sku(value: str) -> bool:
    lower = value.lower()
    if lower in CATEGORY_HEADERS:
        return False
    if value.replace(".", "", 1).isdigit():
        return False
    return True


def read_part_descriptions(path: Path) -> dict[str, tuple[str, str | None, bool]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: dict[str, tuple[str, str | None, bool]] = {}

    for row in ws.iter_rows(min_row=1):
        sku = normalize_text(row[0].value if len(row) > 0 else None)
        if not sku:
            continue
        desc_c = normalize_text(row[2].value if len(row) > 2 else None)
        desc_d = normalize_text(row[3].value if len(row) > 3 else None)
        description = desc_c or desc_d
        if not description:
            continue
        rows[sku] = (description, classify_type(sku), True)

    return rows


def read_book1(path: Path, existing: dict[str, tuple[str, str | None, bool]]) -> dict[str, tuple[str, str | None, bool]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = dict(existing)

    for row in ws.iter_rows(min_row=1):
        sku = normalize_text(row[0].value if len(row) > 0 else None)
        if not sku or not looks_like_book1_sku(sku):
            continue
        description = normalize_text(row[1].value if len(row) > 1 else None)
        if not description:
            continue
        if sku in rows and rows[sku][2]:
            continue
        rows[sku] = (description, classify_type(sku), False)

    return rows


def upsert_rows(catalog_rows: dict[str, tuple[str, str | None, bool]]) -> tuple[int, int]:
    inserted = 0
    updated = 0
    for part_number, (description, product_type, is_active) in catalog_rows.items():
        existing = (
            db.session.query(ProductCatalog).filter_by(part_number=part_number).one_or_none()
        )
        if existing is None:
            db.session.add(
                ProductCatalog(
                    part_number=part_number,
                    description=description,
                    product_type=product_type,
                    is_active=is_active,
                )
            )
            inserted += 1
            continue
        existing.description = description
        existing.product_type = product_type
        existing.is_active = is_active
        updated += 1
    db.session.commit()
    return inserted, updated


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Seed product_catalog from Chip spreadsheets")
    parser.add_argument("--part-descriptions", type=Path, default=DEFAULT_PART_DESCRIPTIONS)
    parser.add_argument("--book1", type=Path, default=DEFAULT_BOOK1)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.part_descriptions.exists():
        raise SystemExit(f"Missing file: {args.part_descriptions}")
    if not args.book1.exists():
        raise SystemExit(f"Missing file: {args.book1}")

    app = create_app()
    with app.app_context():
        active_rows = read_part_descriptions(args.part_descriptions)
        merged_rows = read_book1(args.book1, active_rows)
        inserted, updated = upsert_rows(merged_rows)
        print(
            "Seed complete:"
            f" active_rows={len(active_rows)} total_rows={len(merged_rows)}"
            f" inserted={inserted} updated={updated}"
        )


if __name__ == "__main__":
    main()
